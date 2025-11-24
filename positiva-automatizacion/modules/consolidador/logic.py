"""
Lógica del Consolidador de Servicios Médicos - Anexo 1
VERSIÓN MEJORADA - Búsqueda flexible de hojas
"""
import pandas as pd
import os
import time
from datetime import datetime

# Verificar que pyxlsb está instalado
try:
    from pyxlsb import open_workbook
    PYXLSB_AVAILABLE = True
    print("✓ pyxlsb importado correctamente", flush=True)
except ImportError as e:
    PYXLSB_AVAILABLE = False
    print("⚠️  pyxlsb no disponible - Solo se podrán procesar archivos XLSX", flush=True)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def buscar_hoja_servicios(hojas):
    """
    Busca la hoja de servicios con múltiples criterios
    Prioridad:
    1. Hojas que contengan "TARIFA" Y "SERV"
    2. Hojas que contengan solo "SERV"
    3. Hojas que contengan "RELACION" Y "SERV"
    """
    hoja_target = None
    
    # PRIORIDAD 1: TARIFA + SERV
    for hoja in hojas:
        hoja_upper = hoja.upper()
        if 'TARIFA' in hoja_upper and 'SERV' in hoja_upper:
            print(f"   ✓ Encontrada (TARIFA + SERV): '{hoja}'", flush=True)
            return hoja
    
    # PRIORIDAD 2: Solo SERV (más común)
    for hoja in hojas:
        hoja_upper = hoja.upper()
        if 'SERV' in hoja_upper and 'MEDICO' in hoja_upper:
            print(f"   ✓ Encontrada (SERV + MEDICO): '{hoja}'", flush=True)
            return hoja
    
    # PRIORIDAD 3: RELACION + SERV
    for hoja in hojas:
        hoja_upper = hoja.upper()
        if 'RELACION' in hoja_upper and 'SERV' in hoja_upper:
            print(f"   ✓ Encontrada (RELACION + SERV): '{hoja}'", flush=True)
            return hoja
    
    # PRIORIDAD 4: Solo SERV
    for hoja in hojas:
        hoja_upper = hoja.upper()
        if 'SERV' in hoja_upper:
            print(f"   ✓ Encontrada (solo SERV): '{hoja}'", flush=True)
            return hoja
    
    return None


def leer_archivo_excel(filepath):
    """
    Lee archivo XLSB o XLSX y retorna datos
    
    Returns:
        tuple: (data, hoja_nombre, formato)
    """
    extension = filepath.rsplit('.', 1)[1].lower()
    
    if extension == 'xlsb':
        if not PYXLSB_AVAILABLE:
            raise ImportError("pyxlsb no está instalado. No se pueden procesar archivos XLSB")
        
        print(f"   Formato: XLSB (usando pyxlsb)", flush=True)
        with open_workbook(filepath) as wb:
            hojas = wb.sheets
            print(f"   Hojas totales: {len(hojas)}", flush=True)
            
            # Buscar hoja de servicios
            hoja_target = buscar_hoja_servicios(hojas)
            
            if not hoja_target:
                print(f"   ❌ No se encontró hoja de servicios", flush=True)
                print(f"   Hojas disponibles: {', '.join(hojas)}", flush=True)
                return None, None, 'xlsb'
            
            # Leer datos
            data = []
            with wb.get_sheet(hoja_target) as sheet:
                for row in sheet.rows():
                    data.append([cell.v for cell in row])
            
            return data, hoja_target, 'xlsb'
    
    elif extension == 'xlsx':
        print(f"   Formato: XLSX (usando pandas)", flush=True)
        
        try:
            xls = pd.ExcelFile(filepath)
            hojas = xls.sheet_names
            print(f"   Hojas totales: {len(hojas)}", flush=True)
            
            # Buscar hoja de servicios
            hoja_target = buscar_hoja_servicios(hojas)
            
            if not hoja_target:
                print(f"   ❌ No se encontró hoja de servicios", flush=True)
                print(f"   Hojas disponibles: {', '.join(hojas)}", flush=True)
                return None, None, 'xlsx'
            
            # Leer datos con más opciones para manejar archivos grandes
            df = pd.read_excel(
                filepath, 
                sheet_name=hoja_target, 
                header=None,
                engine='openpyxl'  # Motor más robusto
            )
            data = df.values.tolist()
            
            return data, hoja_target, 'xlsx'
            
        except Exception as e:
            print(f"   ❌ Error leyendo XLSX: {str(e)}", flush=True)
            # Intentar con otro motor
            try:
                print(f"   🔄 Reintentando con motor alternativo...", flush=True)
                xls = pd.ExcelFile(filepath, engine='openpyxl')
                hojas = xls.sheet_names
                hoja_target = buscar_hoja_servicios(hojas)
                
                if not hoja_target:
                    return None, None, 'xlsx'
                
                df = pd.read_excel(filepath, sheet_name=hoja_target, header=None)
                data = df.values.tolist()
                return data, hoja_target, 'xlsx'
            except Exception as e2:
                print(f"   ❌ Error en reintento: {str(e2)}", flush=True)
                raise e2
    
    else:
        return None, None, None


def procesar_anexo1_xlsb(filepath, fecha_acuerdo=None):
    """
    Procesa archivo XLSB o XLSX del Anexo 1
    ESTRUCTURA CORRECTA:
    Col 0: ITEM
    Col 1: CODIGO CUPS
    Col 2: CODIGO HOMOLOGO MANUAL
    Col 3: DESCRIPCION DEL CUPS
    Col 4: TARIFA UNITARIA EN PESOS
    Col 5: MANUAL TARIFARIO
    Col 6: PORCENTAJE MANUAL TARIFARIO
    Col 7: OBSERVACIONES
    """
    inicio = time.time()
    
    print(f"\n{'='*70}", flush=True)
    print(f"🔍 CONSOLIDADOR ANEXO 1 - MULTI-FORMATO", flush=True)
    print(f"{'='*70}", flush=True)
    print(f"📁 Archivo: {os.path.basename(filepath)}", flush=True)
    print(f"📅 Fecha: {fecha_acuerdo if fecha_acuerdo else 'Sin fecha'}", flush=True)
    
    try:
        # Verificar archivo
        if not os.path.exists(filepath):
            return {
                'success': False,
                'error': 'El archivo no existe'
            }
        
        file_size = os.path.getsize(filepath)
        print(f"✓ Tamaño: {file_size:,} bytes ({file_size/1024/1024:.2f} MB)", flush=True)
        
        # Leer archivo (XLSB o XLSX)
        print(f"\n📖 Abriendo archivo...", flush=True)
        data, hoja_target, formato = leer_archivo_excel(filepath)
        
        if data is None:
            return {
                'success': False,
                'error': 'No se encontró la hoja de servicios. Debe contener "SERV" en el nombre.'
            }
        
        print(f"✓ Hoja: '{hoja_target}'", flush=True)
        print(f"✓ Total filas: {len(data):,}", flush=True)
        
        # Procesar sedes y servicios
        print(f"\n🏢 Extrayendo sedes y servicios...", flush=True)
        sedes = []
        sede_actual = None
        servicios_actuales = []
        en_seccion_servicios = False
        
        for idx, row in enumerate(data):
            if not row:
                continue
            
            # Detectar inicio de nueva sede
            es_nueva_sede = False
            for cell in row:
                if cell and isinstance(cell, str):
                    texto = cell.upper()
                    if 'CODIGO DE HABILITACIÓN' in texto or 'CÓDIGO DE HABILITACIÓN' in texto or 'CODIGO DE HABILITACION' in texto:
                        es_nueva_sede = True
                        break
            
            if es_nueva_sede:
                # Guardar sede anterior
                if sede_actual and servicios_actuales:
                    sedes.append({
                        'info': sede_actual,
                        'servicios': servicios_actuales.copy()
                    })
                    print(f"   ✓ {sede_actual['codigo']}: {len(servicios_actuales)} servicios", flush=True)
                
                # Leer info de nueva sede (siguiente fila)
                servicios_actuales = []
                en_seccion_servicios = False
                
                if idx + 1 < len(data):
                    sede_row = data[idx + 1]
                    if sede_row and len(sede_row) > 4:
                        codigo_hab = str(sede_row[2]).strip() if sede_row[2] else ""
                        numero_sede = sede_row[3]
                        
                        # Formatear número de sede
                        if numero_sede is not None:
                            if isinstance(numero_sede, float) and numero_sede.is_integer():
                                numero_sede = int(numero_sede)
                            numero_sede_str = str(numero_sede)
                            if len(numero_sede_str) == 1:
                                numero_sede_str = numero_sede_str.zfill(2)
                        else:
                            numero_sede_str = "01"
                        
                        sede_actual = {
                            'codigo': f"{codigo_hab}-{numero_sede_str}",
                            'codigo_base': codigo_hab,
                            'numero': numero_sede_str,
                            'nombre': str(sede_row[4]).strip() if len(sede_row) > 4 and sede_row[4] else "",
                            'municipio': str(sede_row[1]).strip() if len(sede_row) > 1 and sede_row[1] else ""
                        }
                
                continue
            
            # Detectar encabezado de servicios (ITEM, CODIGO CUPS, etc.)
            if not en_seccion_servicios:
                for cell in row:
                    if cell and isinstance(cell, str):
                        texto = cell.upper()
                        if 'ITEM' in texto and ('CODIGO' in texto or idx + 1 < len(data)):
                            # Verificar si la siguiente columna tiene "CODIGO" o "CUPS"
                            if len(row) > 1 and row[1]:
                                siguiente = str(row[1]).upper()
                                if 'CODIGO' in siguiente or 'CUPS' in siguiente:
                                    en_seccion_servicios = True
                                    break
                
                if en_seccion_servicios:
                    continue
            
            # Extraer servicios
            if en_seccion_servicios and sede_actual and len(row) >= 8:
                try:
                    item = row[0]
                    codigo_cups = row[1]
                    
                    # Validar que tenemos un servicio válido
                    if item is None and codigo_cups is None:
                        continue
                    
                    # El ITEM debe ser un número
                    if item is not None:
                        try:
                            item_num = int(float(item))
                        except:
                            continue
                    
                    # El código CUPS no debe estar vacío
                    if codigo_cups is None or str(codigo_cups).strip() == "":
                        continue
                    
                    codigo_cups_str = str(codigo_cups).strip()
                    
                    # Filtrar encabezados
                    texto_upper = codigo_cups_str.upper()
                    if any(kw in texto_upper for kw in ['CODIGO', 'CUPS', 'DESCRIPCION', 'TARIFA', 'MANUAL']):
                        continue
                    
                    # Extraer servicio
                    servicio = {
                        'item': item,
                        'codigo_cups': codigo_cups_str,
                        'codigo_homologo': str(row[2]).strip() if row[2] else "",
                        'descripcion': str(row[3]).strip() if row[3] else "",
                        'tarifa': row[4] if row[4] is not None else None,
                        'manual': str(row[5]).strip() if row[5] else "",
                        'porcentaje': row[6] if row[6] is not None else None,
                        'observaciones': str(row[7]).strip() if len(row) > 7 and row[7] else ""
                    }
                    
                    servicios_actuales.append(servicio)
                    
                except Exception as e:
                    pass
        
        # Guardar última sede
        if sede_actual and servicios_actuales:
            sedes.append({
                'info': sede_actual,
                'servicios': servicios_actuales.copy()
            })
            print(f"   ✓ {sede_actual['codigo']}: {len(servicios_actuales)} servicios", flush=True)
        
        print(f"\n✓ Sedes procesadas: {len(sedes)}", flush=True)
        
        if not sedes:
            return {
                'success': False,
                'error': 'No se encontraron sedes válidas en el archivo'
            }
        
        # Consolidar
        print(f"\n📦 Consolidando servicios...", flush=True)
        consolidado = []
        vista_previa = []
        
        for sede_data in sedes:
            sede = sede_data['info']
            for servicio in sede_data['servicios']:
                registro = {
                    'codigo_cups': servicio['codigo_cups'],
                    'codigo_homologo_manual': servicio['codigo_homologo'],
                    'descripcion_del_cups': servicio['descripcion'],
                    'tarifa_unitaria_en_pesos': servicio['tarifa'],
                    'manual_tarifario': servicio['manual'],
                    'porcentaje_manual_tarifario': servicio['porcentaje'],
                    'observaciones': servicio['observaciones'],
                    'codigo_de_habilitacion': sede['codigo'],
                    'fecha_acuerdo': fecha_acuerdo if fecha_acuerdo else ""
                }
                consolidado.append(registro)
                
                # Guardar primeros 10 para vista previa
                if len(vista_previa) < 10:
                    vista_previa.append({
                        'cups': servicio['codigo_cups'],
                        'descripcion': servicio['descripcion'][:50] + '...' if len(servicio['descripcion']) > 50 else servicio['descripcion'],
                        'tarifa': servicio['tarifa'],
                        'sede': sede['codigo']
                    })
        
        tiempo_total = round(time.time() - inicio, 2)
        
        print(f"\n{'='*70}", flush=True)
        print(f"✅ PROCESAMIENTO COMPLETADO", flush=True)
        print(f"{'='*70}", flush=True)
        print(f"   Sedes: {len(sedes)}", flush=True)
        print(f"   Servicios: {len(consolidado):,}", flush=True)
        print(f"   Tiempo: {tiempo_total}s", flush=True)
        print(f"\n📋 VISTA PREVIA (primeros 10):", flush=True)
        for i, prev in enumerate(vista_previa, 1):
            print(f"   {i}. CUPS: {prev['cups']} | Sede: {prev['sede']}", flush=True)
            print(f"      {prev['descripcion']}", flush=True)
            print(f"      Tarifa: ${prev['tarifa']:,.2f}" if prev['tarifa'] else "      Tarifa: N/A", flush=True)
        print(f"{'='*70}\n", flush=True)
        
        return {
            'success': True,
            'consolidado': consolidado,
            'vista_previa': vista_previa,
            'total_sedes': len(sedes),
            'total_servicios': len(consolidado),
            'tiempo_ejecucion': tiempo_total
        }
        
    except Exception as e:
        import traceback
        print(f"\n{'='*70}", flush=True)
        print(f"❌ ERROR CRÍTICO", flush=True)
        print(f"{'='*70}", flush=True)
        print(traceback.format_exc(), flush=True)
        print(f"{'='*70}\n", flush=True)
        
        return {
            'success': False,
            'error': f'Error: {str(e)}'
        }


def generar_excel_consolidado(resultado, output_path):
    """
    Genera Excel con formato POSITIVA
    """
    try:
        consolidado = resultado['consolidado']
        
        if not consolidado:
            return False
        
        print(f"\n💾 Generando Excel...", flush=True)
        print(f"   Registros: {len(consolidado):,}", flush=True)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subheader_font = Font(bold=True, size=10, name="Calibri")
        subheader_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        
        # Fila 1: Encabezado principal
        ws.merge_cells('A1:H1')
        cell_a1 = ws['A1']
        cell_a1.value = 'ANEXO 1 PACTADO DEL PRESTADOR'
        cell_a1.fill = header_fill
        cell_a1.font = header_font
        cell_a1.alignment = header_align
        cell_a1.border = thin_border
        
        cell_i1 = ws['I1']
        cell_i1.value = 'INFO ACTA O ACUERDO'
        cell_i1.fill = header_fill
        cell_i1.font = header_font
        cell_i1.alignment = header_align
        cell_i1.border = thin_border
        
        # Fila 2: Nombres de columnas
        columnas = [
            'codigo_cups',
            'codigo_homologo_manual',
            'descripcion_del_cups',
            'tarifa_unitaria_en_pesos',
            'manual_tarifario',
            'porcentaje_manual_tarifario',
            'observaciones',
            'codigo_de_habilitacion',
            'fecha_acuerdo'
        ]
        
        for col_idx, nombre in enumerate(columnas, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = nombre
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = subheader_align
            cell.border = thin_border
        
        # Datos (desde fila 3)
        for row_idx, registro in enumerate(consolidado, 3):
            for col_idx, col_name in enumerate(columnas, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                valor = registro[col_name]
                
                # Manejar valores numéricos
                if col_name in ['tarifa_unitaria_en_pesos', 'porcentaje_manual_tarifario']:
                    if valor is not None and valor != "":
                        try:
                            cell.value = float(valor)
                        except:
                            cell.value = valor
                    else:
                        cell.value = None
                else:
                    cell.value = valor
                
                cell.border = thin_border
                
                # Alineación
                if col_idx in [1, 2, 5, 7, 8, 9]:  # Centrado
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 3:  # Descripción izquierda
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:  # Números derecha
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                
                # Formato numérico
                if col_idx == 4 and cell.value:
                    cell.number_format = '#,##0.00'
                elif col_idx == 6 and cell.value:
                    if isinstance(cell.value, float) and cell.value <= 1:
                        cell.number_format = '0.00%'
                    else:
                        cell.number_format = '#,##0.00'
            
            if row_idx % 100 == 0:
                print(f"   {row_idx - 2:,} registros escritos...", flush=True)
        
        # Anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 55
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 28
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 22
        ws.column_dimensions['I'].width = 18
        
        # Altura de encabezados
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        
        # Guardar
        wb.save(output_path)
        
        size = os.path.getsize(output_path)
        print(f"✓ Guardado: {size:,} bytes ({size/1024:.2f} KB)", flush=True)
        
        return True
        
    except Exception as e:
        import traceback
        print(f"\n❌ Error:", flush=True)
        print(traceback.format_exc(), flush=True)
        return False