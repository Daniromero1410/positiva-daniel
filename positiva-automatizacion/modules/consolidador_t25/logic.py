"""
Lógica de procesamiento para archivos T25
Adaptado del consolidador existente para formato T25
"""
import pandas as pd
import os
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def procesar_archivo_t25(filepath):
    """
    Procesa archivo Excel T25
    
    Args:
        filepath: Ruta del archivo Excel a procesar
        
    Returns:
        dict: Diccionario con resultados del procesamiento
    """
    inicio = time.time()
    
    print(f"\n{'='*70}", flush=True)
    print(f"🔍 CONSOLIDADOR T25", flush=True)
    print(f"{'='*70}", flush=True)
    print(f"📁 Archivo: {os.path.basename(filepath)}", flush=True)
    
    try:
        # Verificar archivo
        if not os.path.exists(filepath):
            return {
                'success': False,
                'error': 'El archivo no existe'
            }
        
        file_size = os.path.getsize(filepath)
        print(f"✓ Tamaño: {file_size:,} bytes ({file_size/1024/1024:.2f} MB)", flush=True)
        
        # Leer archivo Excel
        print(f"\n📖 Leyendo archivo Excel...", flush=True)
        
        # Detectar extensión
        extension = filepath.rsplit('.', 1)[1].lower()
        
        if extension in ['xlsx', 'xls']:
            df = pd.read_excel(filepath, engine='openpyxl')
        elif extension == 'xlsb':
            df = pd.read_excel(filepath, engine='pyxlsb')
        else:
            return {
                'success': False,
                'error': f'Formato no soportado: .{extension}'
            }
        
        print(f"✓ Total filas: {len(df):,}", flush=True)
        print(f"✓ Total columnas: {len(df.columns)}", flush=True)
        
        # Validar estructura básica
        print(f"\n🔍 Validando estructura...", flush=True)
        
        # Aquí puedes agregar validaciones específicas del formato T25
        # Por ahora, procesamos todas las filas
        
        # Procesar datos
        print(f"\n📦 Procesando datos...", flush=True)
        consolidado = []
        
        for idx, row in df.iterrows():
            # Convertir cada fila a diccionario
            registro = row.to_dict()
            consolidado.append(registro)
            
            # Mostrar progreso cada 100 registros
            if (idx + 1) % 100 == 0:
                print(f"   {idx + 1:,} registros procesados...", flush=True)
        
        tiempo_total = round(time.time() - inicio, 2)
        
        print(f"\n{'='*70}", flush=True)
        print(f"✅ PROCESAMIENTO COMPLETADO", flush=True)
        print(f"{'='*70}", flush=True)
        print(f"   Registros: {len(consolidado):,}", flush=True)
        print(f"   Tiempo: {tiempo_total}s", flush=True)
        print(f"{'='*70}\n", flush=True)
        
        return {
            'success': True,
            'consolidado': consolidado,
            'columnas': list(df.columns),
            'total_registros': len(consolidado),
            'tiempo_ejecucion': tiempo_total
        }
        
    except Exception as e:
        import traceback
        print(f"\n{'='*70}", flush=True)
        print(f"❌ ERROR CRÍTICO", flush=True)
        print(f"{'='*70}", flush=True)
        print(traceback.format_exc(), flush=True)
        print(f"{'='*70}\n", flush=True)
        
        return {
            'success': False,
            'error': f'Error: {str(e)}'
        }


def generar_excel_consolidado(resultado, output_path):
    """
    Genera Excel consolidado con formato POSITIVA
    
    Args:
        resultado: Diccionario con datos consolidados
        output_path: Ruta donde guardar el archivo
        
    Returns:
        bool: True si se generó correctamente
    """
    try:
        consolidado = resultado['consolidado']
        columnas = resultado['columnas']
        
        if not consolidado:
            return False
        
        print(f"\n💾 Generando Excel...", flush=True)
        print(f"   Registros: {len(consolidado):,}", flush=True)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidado T25"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        
        # Fila 1: Encabezado principal
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
        cell_header = ws.cell(row=1, column=1)
        cell_header.value = 'CONSOLIDADO T25 - POSITIVA COMPAÑÍA DE SEGUROS'
        cell_header.fill = header_fill
        cell_header.font = header_font
        cell_header.alignment = header_align
        cell_header.border = thin_border
        
        # Fila 2: Nombres de columnas
        for col_idx, nombre_col in enumerate(columnas, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = nombre_col
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = Font(bold=True, size=10, name="Calibri")
            cell.alignment = header_align
            cell.border = thin_border
        
        # Datos (desde fila 3)
        for row_idx, registro in enumerate(consolidado, 3):
            for col_idx, col_name in enumerate(columnas, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                valor = registro.get(col_name)
                
                # Manejar valores None o NaN
                if pd.isna(valor):
                    cell.value = ""
                else:
                    cell.value = valor
                
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            if row_idx % 100 == 0:
                print(f"   {row_idx - 2:,} registros escritos...", flush=True)
        
        # Ajustar ancho de columnas automáticamente
        for col_idx, col_name in enumerate(columnas, 1):
            # Ancho base de 15
            ws.column_dimensions[chr(64 + col_idx)].width = 15
        
        # Altura de encabezados
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        
        # Guardar
        wb.save(output_path)
        
        size = os.path.getsize(output_path)
        print(f"✓ Guardado: {size:,} bytes ({size/1024:.2f} KB)", flush=True)
        
        return True
        
    except Exception as e:
        import traceback
        print(f"\n❌ Error:", flush=True)
        print(traceback.format_exc(), flush=True)
        return False