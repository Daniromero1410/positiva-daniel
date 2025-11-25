"""
Rutas para el módulo Consolidador T25
"""

from flask import Blueprint, render_template, request, jsonify, session, send_file
from werkzeug.utils import secure_filename
import os
import uuid
from datetime import datetime

from .goanywhere import GoAnywhereWebClient
from .consolidator import ConsolidadorT25
from .maestra_manager import MaestraManager
from .stats_manager import StatsManager

consolidador_t25_bp = Blueprint(
    'consolidador_t25',
    __name__,
    template_folder='../../templates/modules/consolidador_t25',
    static_folder='../../static'
)

# Managers globales
maestra_manager = MaestraManager()
stats_manager = StatsManager()

# Almacenamiento de clientes SFTP por sesión
clientes_sftp = {}

# Configuración
UPLOAD_FOLDER = 'data/maestra'
OUTPUT_FOLDER = 'output/consolidador_t25'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


@consolidador_t25_bp.route('/')
def index():
    """Página principal del consolidador T25"""
    return render_template('consolidador_t25.html')


@consolidador_t25_bp.route('/estadisticas')
def obtener_estadisticas():
    """Obtiene estadísticas del módulo"""
    try:
        stats = stats_manager.obtener_estadisticas('consolidador_t25')
        return jsonify(stats), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# GESTIÓN DE MAESTRA
# ============================================================================

@consolidador_t25_bp.route('/maestra/subir', methods=['POST'])
def subir_maestra():
    """Sube archivo de maestra"""
    try:
        if 'archivo' not in request.files:
            return jsonify({'success': False, 'error': 'No se proporcionó archivo'}), 400
        
        archivo = request.files['archivo']
        
        if archivo.filename == '':
            return jsonify({'success': False, 'error': 'Archivo vacío'}), 400
        
        if not archivo.filename.endswith('.xlsb'):
            return jsonify({'success': False, 'error': 'Solo se permiten archivos .xlsb'}), 400
        
        # Guardar archivo
        filename = 'maestra_contratos_vigentes.xlsb'
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        archivo.save(filepath)
        
        # Cargar maestra
        resultado = maestra_manager.cargar_maestra(filepath)
        
        if resultado['success']:
            return jsonify({
                'success': True,
                'total_contratos': resultado['total_contratos'],
                'total_prestadores': resultado['total_prestadores']
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': resultado['error']
            }), 500
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@consolidador_t25_bp.route('/maestra/estado')
def estado_maestra():
    """Verifica estado de la maestra"""
    try:
        if maestra_manager.maestra is None:
            return jsonify({
                'cargada': False,
                'mensaje': 'No hay maestra cargada'
            }), 200
        
        return jsonify({
            'cargada': True,
            'total_contratos': len(maestra_manager.maestra),
            'ultima_carga': maestra_manager.ultima_carga.strftime('%Y-%m-%d %H:%M:%S') if maestra_manager.ultima_carga else None
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# CONEXIÓN A GOANYWHERE
# ============================================================================

@consolidador_t25_bp.route('/goanywhere/conectar', methods=['POST'])
def conectar_goanywhere():
    """Conecta a GoAnywhere"""
    try:
        # Generar ID de sesión único
        session_id = str(uuid.uuid4())
        session['session_id'] = session_id
        
        # Crear cliente
        cliente = GoAnywhereWebClient()
        resultado = cliente.connect()
        
        if resultado['success']:
            clientes_sftp[session_id] = cliente
            return jsonify({
                'success': True,
                'mensaje': 'Conectado a GoAnywhere exitosamente'
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': resultado['error']
            }), 500
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@consolidador_t25_bp.route('/goanywhere/desconectar', methods=['POST'])
def desconectar_goanywhere():
    """Desconecta de GoAnywhere"""
    try:
        session_id = session.get('session_id')
        
        if session_id and session_id in clientes_sftp:
            cliente = clientes_sftp[session_id]
            cliente.disconnect()
            del clientes_sftp[session_id]
        
        return jsonify({
            'success': True,
            'mensaje': 'Desconectado de GoAnywhere'
        }), 200
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@consolidador_t25_bp.route('/goanywhere/estado')
def estado_goanywhere():
    """Verifica estado de conexión a GoAnywhere"""
    try:
        session_id = session.get('session_id')
        
        if not session_id or session_id not in clientes_sftp:
            return jsonify({
                'conectado': False
            }), 200
        
        cliente = clientes_sftp[session_id]
        
        return jsonify({
            'conectado': cliente.is_connected()
        }), 200
    
    except Exception as e:
        return jsonify({
            'conectado': False,
            'error': str(e)
        }), 500


# ============================================================================
# BÚSQUEDA DE CONTRATOS
# ============================================================================

@consolidador_t25_bp.route('/buscar-contrato', methods=['POST'])
def buscar_contrato():
    """Busca contrato en la maestra"""
    try:
        data = request.get_json()
        numero_contrato = data.get('numero_contrato')
        
        if not numero_contrato:
            return jsonify({
                'success': False,
                'error': 'Debe proporcionar número de contrato'
            }), 400
        
        if maestra_manager.maestra is None:
            return jsonify({
                'success': False,
                'error': 'No hay maestra cargada'
            }), 400
        
        contratos = maestra_manager.buscar_contrato(numero_contrato)
        
        if not contratos:
            return jsonify({
                'success': False,
                'error': f'Contrato {numero_contrato} no encontrado en la maestra'
            }), 404
        
        return jsonify({
            'success': True,
            'contratos': contratos
        }), 200
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@consolidador_t25_bp.route('/buscar-contrato/procesar', methods=['POST'])
def procesar_contrato_individual():
    """Procesa un contrato individual"""
    try:
        session_id = session.get('session_id')
        
        if not session_id or session_id not in clientes_sftp:
            return jsonify({
                'success': False,
                'error': 'No hay sesión SFTP activa'
            }), 401
        
        data = request.get_json()
        numero_contrato = data.get('numero_contrato')
        
        if not numero_contrato:
            return jsonify({
                'success': False,
                'error': 'Debe proporcionar número de contrato'
            }), 400
        
        # Buscar información del contrato en la maestra
        contratos = maestra_manager.buscar_contrato(numero_contrato)
        
        if not contratos:
            return jsonify({
                'success': False,
                'error': f'Contrato {numero_contrato} no encontrado en la maestra'
            }), 404
        
        info_contrato = contratos[0]
        
        # Crear consolidador
        cliente = clientes_sftp[session_id]
        consolidador = ConsolidadorT25(cliente)
        
        # Procesar contrato
        print("\n" + "="*70, flush=True)
        print("INICIANDO PROCESAMIENTO DE CONTRATO INDIVIDUAL", flush=True)
        print("="*70 + "\n", flush=True)
        
        resultado = consolidador.procesar_contrato(info_contrato)
        
        print("\n" + "="*70, flush=True)
        print("RESULTADO DEL PROCESAMIENTO", flush=True)
        print(f"Success: {resultado['success']}", flush=True)
        print(f"Anexos descargados: {len(resultado.get('anexos_descargados', []))}", flush=True)
        print(f"Servicios consolidados: {len(resultado.get('servicios_consolidados', []))}", flush=True)
        print(f"Alertas generadas: {len(resultado.get('alertas', []))}", flush=True)
        print("="*70 + "\n", flush=True)
        
        if resultado['success']:
            # Generar Excel consolidado
            archivo_consolidado = generar_excel_consolidado(
                resultado['servicios_consolidados'],
                numero_contrato
            )
            
            print(f"Archivo generado: {archivo_consolidado}", flush=True)
            
            # Registrar estadísticas
            try:
                stats_manager.registrar_proceso(
                    tipo='consolidador_t25_individual',
                    usuario='sistema',
                    archivo=numero_contrato,
                    registros=len(resultado['servicios_consolidados']),
                    exitoso=True
                )
            except:
                pass
            
            return jsonify({
                'success': True,
                'archivo': archivo_consolidado,
                'total_servicios': len(resultado['servicios_consolidados']),
                'total_anexos': len(resultado['anexos_descargados']),
                'alertas': resultado['alertas'],
                'logs': resultado.get('logs', [])
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': resultado.get('error', 'Error desconocido'),
                'alertas': resultado.get('alertas', []),
                'logs': resultado.get('logs', [])
            }), 500
    
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"\nERROR CRITICO EN PROCESAMIENTO:", flush=True)
        print(error_trace, flush=True)
        
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': error_trace
        }), 500


# ============================================================================
# PROCESAMIENTO MASIVO
# ============================================================================

@consolidador_t25_bp.route('/procesar-masivo', methods=['POST'])
def procesar_masivo():
    """Procesa múltiples contratos"""
    try:
        session_id = session.get('session_id')
        
        if not session_id or session_id not in clientes_sftp:
            return jsonify({
                'success': False,
                'error': 'No hay sesión SFTP activa'
            }), 401
        
        if maestra_manager.maestra is None:
            return jsonify({
                'success': False,
                'error': 'No hay maestra cargada'
            }), 400
        
        # Obtener todos los contratos de prestadores de salud
        contratos = maestra_manager.obtener_contratos_prestadores()
        
        cliente = clientes_sftp[session_id]
        consolidador = ConsolidadorT25(cliente)
        
        resultados = []
        servicios_totales = []
        
        print(f"\n{'='*70}")
        print(f"PROCESAMIENTO MASIVO INICIADO")
        print(f"Total de contratos a procesar: {len(contratos)}")
        print(f"{'='*70}\n")
        
        for i, contrato in enumerate(contratos, 1):
            print(f"\nProcesando contrato {i}/{len(contratos)}: {contrato['numero_contrato']}")
            
            resultado = consolidador.procesar_contrato(contrato)
            resultados.append(resultado)
            
            if resultado['success']:
                servicios_totales.extend(resultado['servicios_consolidados'])
                print(f"  Exitoso: {len(resultado['servicios_consolidados'])} servicios")
            else:
                print(f"  Error: {resultado.get('error', 'Desconocido')}")
        
        # Generar archivo consolidado único
        if servicios_totales:
            archivo_consolidado = generar_excel_consolidado(
                servicios_totales,
                'TODOS_LOS_CONTRATOS'
            )
            
            # Registrar estadísticas
            try:
                stats_manager.registrar_proceso(
                    tipo='consolidador_t25_masivo',
                    usuario='sistema',
                    archivo='procesamiento_masivo',
                    registros=len(servicios_totales),
                    exitoso=True
                )
            except:
                pass
            
            return jsonify({
                'success': True,
                'archivo': archivo_consolidado,
                'total_contratos_procesados': len(contratos),
                'total_servicios': len(servicios_totales),
                'total_alertas': len(consolidador.alertas),
                'alertas': consolidador.alertas
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se pudieron procesar contratos',
                'alertas': consolidador.alertas
            }), 500
    
    except Exception as e:
        import traceback
        print(f"\nERROR CRÍTICO EN PROCESAMIENTO MASIVO:")
        print(traceback.format_exc())
        
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# DESCARGA DE ARCHIVOS
# ============================================================================

@consolidador_t25_bp.route('/descargar/<filename>')
def descargar_archivo(filename):
    """Descarga archivo consolidado"""
    try:
        filepath = os.path.join(OUTPUT_FOLDER, filename)
        
        if not os.path.exists(filepath):
            return jsonify({
                'success': False,
                'error': 'Archivo no encontrado'
            }), 404
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@consolidador_t25_bp.route('/alertas')
def obtener_alertas():
    """Obtiene todas las alertas generadas"""
    try:
        session_id = session.get('session_id')
        
        if not session_id or session_id not in clientes_sftp:
            return jsonify({
                'success': False,
                'error': 'No hay sesión activa'
            }), 401
        
        # Aquí podrías implementar almacenamiento persistente de alertas
        return jsonify({
            'success': True,
            'alertas': []
        }), 200
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def generar_excel_consolidado(servicios: list, nombre_base: str) -> str:
    """
    Genera archivo Excel con servicios consolidados
    
    Args:
        servicios: Lista de servicios
        nombre_base: Nombre base para el archivo
        
    Returns:
        Nombre del archivo generado
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Crear DataFrame
        df = pd.DataFrame(servicios)
        
        # Ordenar por contrato y origen
        df = df.sort_values(['numero_contrato_año', 'origen_tarifa'])
        
        # Generar nombre de archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"CONSOLIDADO_{nombre_base}_{timestamp}.xlsx"
        filepath = os.path.join(OUTPUT_FOLDER, filename)
        
        # Crear archivo Excel con formato
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Consolidado')
            
            # Obtener worksheet
            worksheet = writer.sheets['Consolidado']
            
            # Formato de encabezados
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar anchos de columna
            column_widths = {
                'A': 12,  # codigo_cups
                'B': 12,  # codigo_homologo_manual
                'C': 50,  # descripcion_del_cups
                'D': 18,  # tarifa_unitaria_en_pesos
                'E': 20,  # manual_tarifario
                'F': 22,  # porcentaje_manual_tarifario
                'G': 30,  # observaciones
                'H': 20,  # codigo_de_habilitacion
                'I': 15,  # fecha_acuerdo
                'J': 20,  # numero_contrato_año
                'K': 15   # origen_tarifa
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Formato de datos
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        print(f"Archivo Excel generado: {filepath}")
        return filename
    
    except Exception as e:
        print(f"Error generando Excel: {str(e)}")
        import traceback
        print(traceback.format_exc())
        raise