"""
Lógica de negocio para asignación de especialidades médicas
"""
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def asignar_especialidad_multiple(nombre_estudio, es_laboratorio=False):
    """
    Asigna especialidades que pueden ordenar el estudio.
    
    Args:
        nombre_estudio: Nombre del estudio médico
        es_laboratorio: True si es laboratorio, False si es imagen
        
    Returns:
        tuple: (lista de tuplas (codigo, nombre), es_muy_especifico)
    """
    nombre_lower = str(nombre_estudio).lower()
    
    # Especialidades generales que pueden ordenar la mayoría de estudios
    esp_generales = [
        (328, 'MEDICINA GENERAL'),
        (342, 'PEDIATRÍA'),
        (325, 'MEDICINA FAMILIAR'),
        (329, 'MEDICINA INTERNA')
    ]
    
    # ═══════════════════════════════════════════════════════════════════
    # ESTUDIOS MUY ESPECÍFICOS - Solo la especialidad correspondiente
    # ═══════════════════════════════════════════════════════════════════
    
    # CARDIOLOGÍA - Muy específicos
    if any(x in nombre_lower for x in ['electrocardiograma', 'ecocardiograma', 'holter', 
                                         'cateterismo', 'angiografia coronaria']):
        return [(302, 'CARDIOLOGÍA')], True
    
    # GINECOBSTETRICIA - Muy específicos
    if any(x in nombre_lower for x in ['transvaginal', 'histerosonogra', 'colposcopia',
                                         'histeroscopia']):
        return [(320, 'GINECOBSTETRICIA')], True
    
    # GASTROENTEROLOGÍA - Procedimientos endoscópicos
    if any(x in nombre_lower for x in ['esofagogastroduodenoscopia', 'colonoscopia', 
                                         'endoscopia', 'rectosigmoidoscopia']):
        return [(316, 'GASTROENTEROLOGÍA')], True
    
    # NEUROLOGÍA - Muy específicos
    if any(x in nombre_lower for x in ['electromiografía', 'electromiografia', 
                                         'neuroconducción', 'neuroconduccion', 
                                         'potenciales evocados', 'electroencefalograma']):
        return [(332, 'NEUROLOGÍA')], True
    
    # UROLOGÍA - Procedimientos
    if any(x in nombre_lower for x in ['cistoscopia', 'ureteroscopia', 'urodinamia']):
        return [(355, 'UROLOGÍA')], True
    
    # ENDOCRINOLOGÍA - Muy específico
    if any(x in nombre_lower for x in ['osteodensitometria', 'densitometria osea']):
        return [(310, 'ENDOCRINOLOGÍA')], True
    
    # OFTALMOLOGÍA - Muy específicos
    if any(x in nombre_lower for x in ['campo visual', 'golman', 'tomografia optica',
                                         'paquimetria', 'topografia corneal']):
        return [(335, 'OFTALMOLOGÍA')], True
    
    # ═══════════════════════════════════════════════════════════════════
    # ESTUDIOS GENERALES - Múltiples especialidades
    # ═══════════════════════════════════════════════════════════════════
    
    # Cardiovasculares generales
    if any(x in nombre_lower for x in ['stress', 'perfusion miocardica', 'presión arterial']):
        return [(302, 'CARDIOLOGÍA')] + esp_generales, False
    
    # Mamarios
    if any(x in nombre_lower for x in ['mamografia', 'mama ', 'mamaria']):
        return [(320, 'GINECOBSTETRICIA'), (364, 'CIRUGÍA DE MAMA Y TUMORES TEJIDOS BLANDOS'),
                (336, 'ONCOLOGÍA CLÍNICA')] + esp_generales, False
    
    # Ginecológicos
    if any(x in nombre_lower for x in ['pelvica ginecol', 'obstetrica', 'utero', 'ovario']):
        return [(320, 'GINECOBSTETRICIA')] + esp_generales, False
    
    # Gastrointestinales/Abdominales
    if any(x in nombre_lower for x in ['abdomen', 'abdominal', 'higado', 'hepat', 
                                         'páncreas', 'pancrea', 'via biliar', 'vesicula']):
        return [(316, 'GASTROENTEROLOGÍA'), (304, 'CIRUGÍA GENERAL')] + esp_generales, False
    
    # Musculoesqueléticos
    if any(x in nombre_lower for x in ['cadera', 'rodilla', 'extremidades', 'articulacion',
                                         'hombro', 'columna', 'lumbosacra', 'cervical',
                                         'coxo-femoral', 'tobillo', 'codo', 'muñeca',
                                         'mano', 'pie', 'hueso', 'miembro']):
        return [(339, 'ORTOPEDIA Y/O TRAUMATOLOGÍA'), (348, 'REUMATOLOGÍA')] + esp_generales, False
    
    # Respiratorios
    if any(x in nombre_lower for x in ['torax', 'espirom', 'curva flujo', 'pulmon',
                                         'respiratorio', 'bronquio']):
        return [(331, 'NEUMOLOGÍA')] + esp_generales, False
    
    # Endocrinos
    if any(x in nombre_lower for x in ['tiroides', 'paratiroides']):
        return [(310, 'ENDOCRINOLOGÍA')] + esp_generales, False
    
    # Neurológicos generales
    if any(x in nombre_lower for x in ['cerebro', 'craneo', 'encefal']):
        return [(332, 'NEUROLOGÍA')] + esp_generales, False
    
    # Urológicos
    if any(x in nombre_lower for x in ['vias urinarias', 'renal', 'riñon', 'prostata',
                                         'vejiga', 'testicular']):
        return [(355, 'UROLOGÍA')] + esp_generales, False
    
    # Vasculares
    if any(x in nombre_lower for x in ['doppler', 'vasos venosos', 'arterial', 'venoso',
                                         'vascular', 'carotidas', 'duplex']):
        return [(372, 'CIRUGÍA VASCULAR'), (302, 'CARDIOLOGÍA')] + esp_generales, False
    
    # ORL
    if any(x in nombre_lower for x in ['senos paranasales', 'nariz', 'oido', 'laringe', 'faringe']):
        return [(340, 'OTORRINOLARINGOLOGÍA')] + esp_generales, False
    
    # Cuello
    if 'cuello' in nombre_lower:
        return [(340, 'OTORRINOLARINGOLOGÍA'), (362, 'CIRUGÍA DE CABEZA Y CUELLO')] + esp_generales, False
    
    # ═══════════════════════════════════════════════════════════════════
    # LABORATORIOS CLÍNICOS
    # ═══════════════════════════════════════════════════════════════════
    
    if es_laboratorio:
        # Hormonas
        if any(x in nombre_lower for x in ['hormona', 'tsh', 't3', 't4', 'tiroides']):
            return [(310, 'ENDOCRINOLOGÍA')] + esp_generales, False
        
        # Hematología
        if any(x in nombre_lower for x in ['hemograma', 'leucocitos', 'plaquetas', 'eritrocitos',
                                             'hematocrito', 'hemoglobina', 'coagulacion', 'protrombina',
                                             'ptt', 'fibrinogeno', 'dimero']):
            return [(321, 'HEMATOLOGÍA')] + esp_generales, False
        
        # Función renal
        if any(x in nombre_lower for x in ['creatinina', 'urea', 'nitrogeno ureico', 'bun']):
            return [(330, 'NEFROLOGÍA')] + esp_generales, False
        
        # Infecciosos/Inmunológicos
        if any(x in nombre_lower for x in ['anticuerpos', 'antigeno', 'serologia', 'hepatitis',
                                             'vih', 'hiv', 'vdrl', 'toxoplasma', 'rubeola',
                                             'citomegalovirus', 'herpes', 'iga', 'igg', 'igm']):
            return [(323, 'INFECTOLOGÍA')] + esp_generales, False
        
        # Marcadores tumorales
        if any(x in nombre_lower for x in ['alfa feto', 'cea', 'ca 125', 'ca 19-9', 'ca 15-3',
                                             'psa', 'especifico de prostata']):
            return [(336, 'ONCOLOGÍA CLÍNICA'), (355, 'UROLOGÍA')] + esp_generales, False
        
        # Química general
        if any(x in nombre_lower for x in ['glucosa', 'colesterol', 'trigliceridos',
                                             'transaminasas', 'got', 'gpt', 'alt', 'ast',
                                             'fosfatasa', 'bilirrubina', 'albumina', 'proteinas',
                                             'electrolitos', 'sodio', 'potasio', 'calcio',
                                             'acido urico', 'amilasa', 'lipasa']):
            return esp_generales, False
    
    # Estudios muy generales
    if any(x in nombre_lower for x in ['tejido blando', 'citologia', 'biopsia', 'especimen',
                                         'parcial de orina', 'coprologic', 'hemoclasificacion']):
        return esp_generales, False
    
    # Por defecto: estudios generales
    return esp_generales, False


def procesar_excel(file_path):
    """
    Procesa el archivo Excel y asigna especialidades
    
    Args:
        file_path: Ruta del archivo Excel a procesar
        
    Returns:
        dict: Diccionario con resultados del procesamiento
    """
    inicio = datetime.now()
    
    try:
        # Leer archivo Excel
        xls = pd.ExcelFile(file_path)
        
        resultados = {
            'success': True,
            'fecha_proceso': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'hojas_procesadas': [],
            'total_estudios': 0,
            'estudios_especificos': 0,
            'estudios_generales': 0,
            'datos_imagenes': None,
            'datos_laboratorio': None,
            'errores': []
        }
        
        # Procesar hoja "Imagenes"
        if 'Imagenes' in xls.sheet_names:
            df_imagenes = pd.read_excel(file_path, sheet_name='Imagenes')
            resultados_img = []
            
            for idx, row in df_imagenes.iterrows():
                codigo_estudio = row['Servicio_Principal']
                nombre_estudio = row['Nombre_Servicio_Principal_ajust']
                
                especialidades_list, es_especifico = asignar_especialidad_multiple(nombre_estudio, False)
                
                if es_especifico:
                    resultados['estudios_especificos'] += 1
                else:
                    resultados['estudios_generales'] += 1
                
                codigos = [str(int(esp[0])) for esp in especialidades_list]
                nombres = [esp[1] for esp in especialidades_list]
                especialidad_str = ' | '.join([f"{cod} - {nom}" for cod, nom in zip(codigos, nombres)])
                
                resultados_img.append({
                    'Servicio_Principal': codigo_estudio,
                    'Nombre_Servicio_Principal_ajust': nombre_estudio,
                    'PLAN 1': row.get('PLAN 1'),
                    'PLAN 2': row.get('PLAN 2'),
                    'PLAN 3': row.get('PLAN 3'),
                    'codigo especialidad principal': codigos[0] if len(codigos) == 1 else ', '.join(codigos),
                    'Nombre Especialidad principal': especialidad_str
                })
            
            resultados['datos_imagenes'] = pd.DataFrame(resultados_img)
            resultados['hojas_procesadas'].append('Imagenes')
            resultados['total_estudios'] += len(resultados_img)
        
        # Procesar hoja "Laboratorio Clinico"
        if 'Laboratorio Clinico' in xls.sheet_names:
            df_laboratorio = pd.read_excel(file_path, sheet_name='Laboratorio Clinico')
            resultados_lab = []
            
            for idx, row in df_laboratorio.iterrows():
                codigo_estudio = row['Servicio_Principal']
                nombre_estudio = row['Nombre_Servicio_Principal_ajust']
                
                especialidades_list, es_especifico = asignar_especialidad_multiple(nombre_estudio, True)
                
                if es_especifico:
                    resultados['estudios_especificos'] += 1
                else:
                    resultados['estudios_generales'] += 1
                
                codigos = [str(int(esp[0])) for esp in especialidades_list]
                nombres = [esp[1] for esp in especialidades_list]
                especialidad_str = ' | '.join([f"{cod} - {nom}" for cod, nom in zip(codigos, nombres)])
                
                resultados_lab.append({
                    'Servicio_Principal': codigo_estudio,
                    'Nombre_Servicio_Principal_ajust': nombre_estudio,
                    'can': row.get('can'),
                    'PLAN 1': row.get('PLAN 1'),
                    'PLAN 2': row.get('PLAN 2'),
                    'PLAN 3': row.get('PLAN 3'),
                    'especialidad principal': especialidad_str
                })
            
            resultados['datos_laboratorio'] = pd.DataFrame(resultados_lab)
            resultados['hojas_procesadas'].append('Laboratorio Clinico')
            resultados['total_estudios'] += len(resultados_lab)
        
        # Calcular tiempo de ejecución
        fin = datetime.now()
        tiempo_ejecucion = (fin - inicio).total_seconds()
        
        # Registrar proceso exitoso en estadísticas
        from utils.stats import stats_manager
        stats_manager.registrar_proceso(
            modulo='especialidades',
            archivo_nombre=os.path.basename(file_path),
            total_registros=resultados['total_estudios'],
            estudios_especificos=resultados['estudios_especificos'],
            estudios_generales=resultados['estudios_generales'],
            exito=True,
            tiempo_ejecucion=tiempo_ejecucion
        )
        
        resultados['tiempo_ejecucion'] = round(tiempo_ejecucion, 2)
        
        return resultados
        
    except Exception as e:
        # Calcular tiempo hasta el error
        fin = datetime.now()
        tiempo_ejecucion = (fin - inicio).total_seconds()
        
        # Registrar proceso fallido
        from utils.stats import stats_manager
        stats_manager.registrar_proceso(
            modulo='especialidades',
            archivo_nombre=os.path.basename(file_path),
            total_registros=0,
            exito=False,
            tiempo_ejecucion=tiempo_ejecucion
        )
        
        return {
            'success': False,
            'error': str(e),
            'errores': [str(e)]
        }


def generar_excel_resultado(resultados, output_path):
    """
    Genera el archivo Excel con los resultados
    
    Args:
        resultados: Diccionario con los resultados del procesamiento
        output_path: Ruta donde guardar el archivo
        
    Returns:
        bool: True si se generó correctamente
    """
    try:
        # Crear Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Escribir hojas
            if resultados['datos_imagenes'] is not None:
                resultados['datos_imagenes'].to_excel(writer, sheet_name='Imagenes', index=False)
            
            if resultados['datos_laboratorio'] is not None:
                resultados['datos_laboratorio'].to_excel(writer, sheet_name='Laboratorio Clinico', index=False)
            
            # Hoja de resumen
            resumen_data = []
            resumen_data.append(['REPORTE DE ASIGNACIÓN AUTOMATIZADA'])
            resumen_data.append([])
            resumen_data.append(['Fecha de procesamiento:', resultados['fecha_proceso']])
            resumen_data.append(['Total de estudios:', resultados['total_estudios']])
            resumen_data.append(['Estudios muy específicos (1 especialidad):', resultados['estudios_especificos']])
            resumen_data.append(['Estudios generales (múltiples especialidades):', resultados['estudios_generales']])
            resumen_data.append(['Tiempo de ejecución:', f"{resultados.get('tiempo_ejecucion', 0)} segundos"])
            resumen_data.append([])
            resumen_data.append(['ESPECIALIDADES GENERALES INCLUIDAS:'])
            resumen_data.append(['328 - MEDICINA GENERAL'])
            resumen_data.append(['342 - PEDIATRÍA'])
            resumen_data.append(['325 - MEDICINA FAMILIAR'])
            resumen_data.append(['329 - MEDICINA INTERNA'])
            
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False, header=False)
        
        # Aplicar formato
        wb = load_workbook(output_path)
        
        header_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formatear hojas de datos
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Imagenes', 'Laboratorio Clinico']:
                ws = wb[sheet_name]
                
                # Ajustar anchos
                ws.column_dimensions['A'].width = 18
                ws.column_dimensions['B'].width = 65
                ws.column_dimensions['C'].width = 10
                ws.column_dimensions['D'].width = 10
                ws.column_dimensions['E'].width = 10
                ws.column_dimensions['F'].width = 30
                ws.column_dimensions['G'].width = 120
                
                # Formatear encabezados
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                
                # Formatear celdas de datos
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                ws.freeze_panes = 'A2'
        
        wb.save(output_path)
        return True
        
    except Exception as e:
        print(f"Error al generar Excel: {e}")
        return False